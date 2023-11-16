from openpyxl import Workbook

wb = Workbook()
ws1 = wb.active
ws1.title = "Temperatura"
ws1.sheet_properties.tabColor = 'f07a7a'
ws1.column_dimensions['A'].width = 18
ws1.column_dimensions['B'].width = 18

ws2 = wb.create_sheet("Presión")
ws2.sheet_properties.tabColor = '87c8e0'
ws2.column_dimensions['A'].width = 18
ws2.column_dimensions['B'].width = 18

ws3 = wb.create_sheet("Radiación")
ws3.sheet_properties.tabColor = '8ce690'
ws3.column_dimensions['A'].width = 18
ws3.column_dimensions['B'].width = 18

ws4 = wb.create_sheet("Luz")
ws4.sheet_properties.tabColor = 'eff0a8'
ws4.column_dimensions['A'].width = 18
ws4.column_dimensions['B'].width = 18

ws1['B1'] = 'TIEMPO'
ws1['A1'] = 'LECTURAS'

ws2['B1'] = 'TIEMPO'
ws2['A1'] = 'LECTURAS'

ws3['B1'] = 'TIEMPO'
ws3['A1'] = 'LECTURAS'

ws4['B1'] = 'TIEMPO'
ws4['A1'] = 'LECTURAS'


def exportar(val):
    nom = None
    for a in range(2):
        for cell in range((len(tab1.xs))):
            if a == 0:
                nom = 'A' + str(cell + 2)
                ws1[nom] = tab1.ys[cell]
                ws2[nom] = tab2.ys[cell]
                ws3[nom] = tab3.ys[cell]
                ws4[nom] = tab4.ys[cell]
            if a == 1:
                nom = 'B' + str(cell + 2)
                ws1[nom] = tab1.xs[cell]
                ws2[nom] = tab2.xs[cell]
                ws3[nom] = tab3.xs[cell]
                ws4[nom] = tab4.xs[cell]
