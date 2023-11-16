from datetime import datetime as dt
from atexit import register
import matplotlib.pyplot as plt
import matplotlib.animation as animation
from matplotlib.widgets import Button
from matplotlib import rcParams
import atexit
from openpyxl import Workbook

from sx126x import sx126x


def cm_a_inch(valor):
    return valor / 2.54


node = sx126x(serial_num="COM6", freq=433, addr=0, power=22, rssi=True)

# Creamos la figura

rcParams['toolbar'] = 'None'
fig = plt.figure(figsize=(cm_a_inch(34), cm_a_inch(17)))

fig.canvas.manager.set_window_title('Galileo Nano 2022')

archivo = open('exportación.txt', 'w')
archivo.write("")
archivo.close()

archivo_hora = open('hora.txt', 'w')
archivo_hora.write("")
archivo_hora.close()

ax = fig.add_subplot(2, 2, 1)
bx = fig.add_subplot(2, 2, 2)
cx = fig.add_subplot(2, 2, 3)
dx = fig.add_subplot(2, 2, 4)

plt.subplots_adjust(top=0.934,
                    bottom=0.133,
                    left=0.044,
                    right=0.971,
                    hspace=0.217,
                    wspace=0.092)


class Argumentos:
    xs = []
    ys = []

    def __init__(self, xs, ys):
        self.xs = xs
        self.ys = ys


tab1 = Argumentos
tab2 = Argumentos
tab3 = Argumentos
tab4 = Argumentos

# TODO LO DE EXCEL

wb = Workbook()
ws1 = wb.active
ws1.title = "Temperatura"
ws1.sheet_properties.tabColor = 'f07a7a'
ws1.column_dimensions['A'].width = 18
ws1.column_dimensions['B'].width = 18

ws2 = wb.create_sheet("Presión")
ws2.sheet_properties.tabColor = '87c8e0'
ws2.column_dimensions['A'].width = 18
ws2.column_dimensions['B'].width = 18

ws3 = wb.create_sheet("Radiación")
ws3.sheet_properties.tabColor = '8ce690'
ws3.column_dimensions['A'].width = 18
ws3.column_dimensions['B'].width = 18

ws4 = wb.create_sheet("Luz")
ws4.sheet_properties.tabColor = 'eff0a8'
ws4.column_dimensions['A'].width = 18
ws4.column_dimensions['B'].width = 18

ws1['B1'] = 'TIEMPO'
ws1['A1'] = 'LECTURAS'

ws2['B1'] = 'TIEMPO'
ws2['A1'] = 'LECTURAS'

ws3['B1'] = 'TIEMPO'
ws3['A1'] = 'LECTURAS'

ws4['B1'] = 'TIEMPO'
ws4['A1'] = 'LECTURAS'

# Función utilizada por al función de animate

a = 0
b = 0
temp_aa = ""
temp_bb = ""
temp_cc = ""
temp_dd = ""

archivo = open('exportación.txt', 'a')
archivo_hora = open('hora.txt', 'a')

def Animation(i, abc, bbc, cbc, dbc, a, b, temp_a, temp_b, temp_c, temp_d, archivo_datos, dato_hora):
    hora = dt.now().strftime("%H:%M:%S")

    datos_recibidos = str(node.receive())
    print(datos_recibidos)

    dato_hora.write(hora + '\n')

    if datos_recibidos[:3] == 'AAA':
        for z in datos_recibidos:

            a = a + 1
            if a < 5:
                continue
            elif a >= 5:
                if z == " ":
                    b = b + 1
                    continue
                elif a != " ":
                    if b == 0:
                        temp_a = temp_a + str(z)
                    if b == 1:
                        temp_b = temp_b + str(z)
                    if b == 2:
                        temp_c = temp_c + str(z)
                    if b == 3:
                        temp_d = temp_d + str(z)
                    if b > 3:
                        break

    archivo_datos.write(temp_a+" "+temp_b+" "+temp_c+" "+temp_d + '\n')
    # TEMPERATURA

    plt.subplot(2, 2, 1)

    # Añandiendo los datos
    abc.xs.append(str(hora))
    abc.ys.append(temp_a)
    # Redibujar la grafica
    ax.clear()
    # if not paused:
    ax.plot(abc.ys[-400:], )
    # Cambiar el formato de la tabla
    plt.title('TEMPERATURA')
    # plt.axis('off')

    # PRESIÓN

    plt.subplot(2, 2, 2)
    # Datos random
    # Añandiendo los datos
    bbc.xs.append(hora)
    bbc.ys.append(temp_b)
    # Redibujar la grafica
    bx.clear()
    # if not paused:
    bx.plot(bbc.ys[-400:], )
    # Cambiar el formato de la tabla
    plt.title('PRESIÓN')
    # plt.axis('off')

    # RADIACIÓN

    plt.subplot(2, 2, 3)
    # Datos random
    # Añandiendo los datos
    cbc.xs.append(hora)
    cbc.ys.append(temp_c)
    # Redibujar la grafica
    cx.clear()
    # if not paused:
    cx.plot(cbc.ys[-400:], )
    # Cambiar el formato de la tabla
    plt.title('RADIACIÓN')
    # plt.axis('off')

    # LUZ

    plt.subplot(2, 2, 4)
    # Datos random
    # Añandiendo los datos
    dbc.xs.append(hora)
    dbc.ys.append(temp_d)
    # Redibujar la grafica
    dx.clear()
    # if not paused:
    dx.plot(dbc.ys[-400:], )
    # Cambiar el formato de la tabla
    plt.title('LUZ')
    # plt.axis('off')


def empezar(val):
    tab1.xs = []
    tab1.ys = []

    tab2.xs = []
    tab2.ys = []

    tab3.xs = []
    tab3.ys = []

    tab4.xs = []
    tab4.ys = []


def exportar(val):
    nom = None
    for a in range(2):
        for cell in range((len(tab1.xs))):
            if a == 0:
                nom = 'A' + str(cell + 2)
                ws1[nom] = tab1.ys[cell]
                ws2[nom] = tab2.ys[cell]
                ws3[nom] = tab3.ys[cell]
                ws4[nom] = tab4.ys[cell]
            if a == 1:
                nom = 'B' + str(cell + 2)
                ws1[nom] = tab1.xs[cell]
                ws2[nom] = tab2.xs[cell]
                ws3[nom] = tab3.xs[cell]
                ws4[nom] = tab4.xs[cell]


"""
░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░
░░░░░░░░░░░▄▀▄▀▀▀▀▄▀▄░░░░░░░░░░░░░░░░░░
░░░░░░░░░░░█░░░░░░░░▀▄░░░░░░▄░░░░░░░░░░
░░░░░░░░░░█░░▀░░▀░░░░░▀▄▄░░█░█░░░░░░░░░
░░░░░░░░░░█░▄░█▀░▄░░░░░░░▀▀░░█░░░░░░░░░
░░░░░░░░░░█░░▀▀▀▀░░░░░░░░░░░░█░░░░░░░░░
░░░░░░░░░░█░░░░░░░░░░░░░░░░░░█░░░░░░░░░
░░░░░░░░░░█░░░░░░░░░░░░░░░░░░█░░░░░░░░░
░░░░░░░░░░░█░░▄▄░░▄▄▄▄░░▄▄░░█░░░░░░░░░░
░░░░░░░░░░░█░▄▀█░▄▀░░█░▄▀█░▄▀░░░░░░░░░░
░░░░░░░░░░░░▀░░░▀░░░░░▀░░░▀░░░░░░░░░░░░
░░░░░░Code made by Zana with <3░░░░░░░░
░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░
"""

axes1 = plt.axes([0.64, 0.008, 0.15, 0.065])
bempezar = Button(axes1, 'Empezar recogida')
bempezar.on_clicked(empezar)

axes2 = plt.axes([0.82, 0.008, 0.15, 0.065])
bacabar = Button(axes2, 'Exportar XLSX')
bacabar.on_clicked(exportar)

ani = animation.FuncAnimation(fig=fig,
                              func=Animation,
                              fargs=(tab1, tab2, tab3, tab4, a, b, temp_aa, temp_bb, temp_cc, temp_dd, archivo, archivo_hora),
                              interval=1250)
plt.show()
atexit.register(wb.save, 'Datos.xlsx')
atexit.register(func=archivo.close())
atexit.register(func=archivo_hora.close())